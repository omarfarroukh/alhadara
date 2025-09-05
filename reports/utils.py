# reports/utils.py
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def style_header_row(sheet, row_number=1):
    """Applies bold font and center alignment to all cells in a given row."""
    header_font = Font(bold=True, name='Calibri', size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # This function is safe because it iterates by row, where merged cells are handled correctly.
    for cell in sheet[row_number]:
        cell.font = header_font
        cell.alignment = center_alignment

def adjust_column_widths(sheet):
    """
    Adjusts column widths based on the longest cell content.
    This version is safe for sheets with merged cells.
    """
    # Iterate through columns by their index (1, 2, 3...) instead of column objects
    for col_idx in range(1, sheet.max_column + 1):
        max_length = 0
        # Safely get the column letter from its index
        column_letter = get_column_letter(col_idx)
        
        # Iterate through all cells in this column
        for cell in sheet[column_letter]:
            # We only care about cells that actually have a value.
            # Merged cell placeholders have a value of None.
            if cell.value is not None:
                try:
                    # Check the length of the cell's content
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    # Pass if cell content has no length
                    pass
                    
        # Add a little padding to the width for aesthetics
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width