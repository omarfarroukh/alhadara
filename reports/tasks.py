import openpyxl
from io import BytesIO
from django.core.files.base import ContentFile
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from decimal import Decimal

from .models import Report
from core.models import Transaction, User, Profile, FileStorage
from complaints.models import Complaint
from courses.models import Enrollment, Booking, Course, Hall, ScheduleSlot
from feedback.models import Feedback
from quiz.models import QuizAttempt
from lessons.models import Attendance, HomeworkGrade
from .utils import style_header_row, adjust_column_widths
from core.services import upload_to_telegram

def _save_report_file(report, workbook, filename_prefix):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    file_name = f"{filename_prefix}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    content_file = ContentFile(buffer.getvalue(), name=file_name)

    tg_result = upload_to_telegram(content_file)
    
    file_storage = FileStorage.objects.create(
        telegram_file_id=tg_result['file_id'],
        telegram_download_link=tg_result['download_link'],
        uploaded_by=report.requested_by
    )
    
    report.file_storage = file_storage
    report.status = 'completed'
    report.completed_at = timezone.now()
    report.save()

def _handle_task_failure(report_id, error):
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'failed'
        report.error_message = str(error)
        report.completed_at = timezone.now()
        report.save()
    except Report.DoesNotExist:
        print(f"Failed to log error for non-existent report_id: {report_id}")

# --- REPORT TASKS ---

def generate_financial_report(report_id, start_date, end_date):
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'processing'; report.save()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Financial Report"
        date_filter = Q(created_at__date__gte=start_date) & Q(created_at__date__lte=end_date)
        
        enrollment_revenue = Transaction.objects.filter(date_filter, transaction_type='course_payment', status='completed').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        booking_revenue = Transaction.objects.filter(date_filter, transaction_type='booking_payment', status='completed').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        total_revenue = enrollment_revenue + booking_revenue
        course_refunds = Transaction.objects.filter(date_filter, transaction_type='course_refund', status='completed').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        booking_refunds = Transaction.objects.filter(date_filter, transaction_type='booking_refund', status='completed').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        total_refunds = course_refunds + booking_refunds
        deposits = Transaction.objects.filter(date_filter, transaction_type='deposit', status='completed').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        withdrawals = Transaction.objects.filter(date_filter, transaction_type='withdrawal', status='completed').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        sheet.append([f"Financial Report ({start_date} to {end_date})"]); sheet.merge_cells('A1:B1'); sheet.append([])
        headers = ['Metric', 'Value (SYP)']; data = [('Total Revenue', total_revenue), ('  - Course Enrollment Revenue', enrollment_revenue), ('  - Hall Booking Revenue', booking_revenue), ('Total Refunds', total_refunds), ('  - Course Refunds', course_refunds), ('  - Hall Booking Refunds', booking_refunds), ('Net Revenue (Revenue - Refunds)', total_revenue - total_refunds), ('Total Deposits', deposits), ('Total Withdrawals', withdrawals)]
        sheet.append(headers)
        for row in data: sheet.append(row)
        style_header_row(sheet, row_number=3); adjust_column_widths(sheet)
        _save_report_file(report, wb, f"financial_{start_date}_{end_date}")
    except Exception as e: _handle_task_failure(report_id, e)

def generate_statistical_report(report_id, start_date, end_date):
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'processing'
        report.save()

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Statistical Report"

        # --- Filters ---
        # Filter for querying the Enrollment model directly
        enrollment_date_filter_direct = Q(enrollment_date__date__gte=start_date) & Q(enrollment_date__date__lte=end_date)
        
        # Filter for querying from Course -> Enrollment relationship
        enrollment_date_filter_related = Q(enrollments__enrollment_date__date__gte=start_date) & Q(enrollments__enrollment_date__date__lte=end_date)
        
        # Filter for querying the Booking model directly
        booking_date_filter_direct = Q(date__gte=start_date) & Q(date__lte=end_date)

        # CORRECTED: Filter for querying from Hall -> Booking relationship
        booking_date_filter_related = Q(bookings__date__gte=start_date) & Q(bookings__date__lte=end_date)

        # --- Data Aggregation ---
        active_enrollments = Enrollment.objects.filter(enrollment_date_filter_direct, status='active').count()
        cancelled_enrollments = Enrollment.objects.filter(enrollment_date_filter_direct, status='cancelled').count()
        total_bookings = Booking.objects.filter(booking_date_filter_direct, status='approved').count()
        
        top_courses = Course.objects.annotate(
            num_enrollments=Count('enrollments', filter=enrollment_date_filter_related)
        ).order_by('-num_enrollments')[:5]
        
        bottom_courses = Course.objects.annotate(
            num_enrollments=Count('enrollments', filter=enrollment_date_filter_related)
        ).order_by('num_enrollments')[:5]
        
        # CORRECTED: Use the related filter for the annotation
        top_halls = Hall.objects.annotate(
            num_bookings=Count('bookings', filter=booking_date_filter_related)
        ).order_by('-num_bookings')[:3]
        
        bottom_halls = Hall.objects.annotate(
            num_bookings=Count('bookings', filter=booking_date_filter_related)
        ).order_by('num_bookings')[:3]

        # --- Build Excel Sheet (No changes needed below this line) ---
        sheet.append([f"Statistical Report ({start_date} to {end_date})"])
        sheet.merge_cells('A1:B1')
        sheet.append([]) # Spacer

        sheet.append(['Metric', 'Value'])
        style_header_row(sheet, row_number=3)
        
        data = [
            ('Active Enrollments', active_enrollments),
            ('Cancelled Enrollments', cancelled_enrollments),
            ('Total Hall Bookings', total_bookings),
        ]
        for row in data: sheet.append(row)
        
        sheet.append([]); sheet.append(['Top 5 Enrolled Courses'])
        for course in top_courses: sheet.append([f"  - {course.title}", course.num_enrollments])

        sheet.append([]); sheet.append(['Bottom 5 Enrolled Courses'])
        for course in bottom_courses: sheet.append([f"  - {course.title}", course.num_enrollments])
        
        sheet.append([]); sheet.append(['Top 3 Booked Halls'])
        for hall in top_halls: sheet.append([f"  - {hall.name}", hall.num_bookings])

        sheet.append([]); sheet.append(['Bottom 3 Booked Halls'])
        for hall in bottom_halls: sheet.append([f"  - {hall.name}", hall.num_bookings])
        
        adjust_column_widths(sheet)

        _save_report_file(report, wb, f"statistical_{start_date}_{end_date}")

    except Exception as e:
        _handle_task_failure(report_id, e)

def generate_feedback_report(report_id, start_date, end_date):
    try:
        report = Report.objects.get(id=report_id); report.status = 'processing'; report.save()
        wb = openpyxl.Workbook(); summary_sheet = wb.active; summary_sheet.title = "Feedback Summary"
        date_filter = Q(created_at__date__gte=start_date) & Q(created_at__date__lte=end_date)
        feedbacks = Feedback.objects.filter(date_filter)
        summary_sheet.append([f"Feedback Summary ({start_date} to {end_date})"]); summary_sheet.merge_cells('A1:B1'); summary_sheet.append([])
        
        if not feedbacks.exists():
            summary_sheet.append(["No feedback submitted in this period."])
        else:
            averages = feedbacks.aggregate(avg_teacher=Avg('teacher_rating'), avg_material=Avg('material_rating'), avg_facilities=Avg('facilities_rating'), avg_app=Avg('app_rating'))
            overall_avg = (averages['avg_teacher'] + averages['avg_material'] + averages['avg_facilities'] + averages['avg_app']) / 4
            summary_sheet.append(['Metric', 'Average Rating (out of 100)']); style_header_row(summary_sheet, 3)
            summary_data = [('Average Teacher Rating', f"{averages['avg_teacher']:.2f}"), ('Average Material Rating', f"{averages['avg_material']:.2f}"), ('Average Facilities Rating', f"{averages['avg_facilities']:.2f}"), ('Average App Rating', f"{averages['avg_app']:.2f}"), ('Overall Average Rating', f"{overall_avg:.2f}"), ('Total Feedback Submissions', feedbacks.count())]
            for row in summary_data: summary_sheet.append(row)
        adjust_column_widths(summary_sheet)
        
        raw_sheet = wb.create_sheet(title="All Feedback Data")
        raw_sheet.append(['Date', 'Student', 'Course', 'Teacher', 'Teacher Rating', 'Material Rating', 'Facilities Rating', 'App Rating', 'Notes']); style_header_row(raw_sheet)
        for fb in feedbacks.select_related('student', 'scheduleslot__course', 'scheduleslot__teacher'):
            raw_sheet.append([fb.created_at.strftime('%Y-%m-%d'), fb.student.get_full_name(), fb.scheduleslot.course.title, fb.scheduleslot.teacher.get_full_name(), fb.teacher_rating, fb.material_rating, fb.facilities_rating, fb.app_rating, fb.notes])
        adjust_column_widths(raw_sheet)
        _save_report_file(report, wb, f"feedback_{start_date}_{end_date}")
    except Exception as e: _handle_task_failure(report_id, e)

def generate_complaints_report(report_id, start_date, end_date):
    try:
        report = Report.objects.get(id=report_id)
        report.status = 'processing'
        report.save()

        # CORRECTED LINE: Changed 'openpypyxl' to 'openpyxl'
        wb = openpyxl.Workbook()
        
        summary_sheet = wb.active
        summary_sheet.title = "Complaints Summary"
        date_filter = Q(created_at__date__gte=start_date) & Q(created_at__date__lte=end_date)
        
        complaints = Complaint.objects.filter(date_filter).select_related(
            'student', 
            'enrollment__course', 
            'enrollment__schedule_slot__teacher'
        )

        summary_sheet.append([f"Complaints Summary ({start_date} to {end_date})"])
        summary_sheet.merge_cells('A1:B1')
        summary_sheet.append([])

        if not complaints.exists():
            summary_sheet.append(["No complaints submitted in this period."])
        else:
            status_breakdown = complaints.values('status').annotate(count=Count('id'))
            priority_breakdown = complaints.values('priority').annotate(count=Count('id'))
            type_breakdown = complaints.values('type').annotate(count=Count('id'))

            summary_sheet.append(['Total Complaints', complaints.count()]); summary_sheet.append([])
            
            summary_sheet.append(['By Status']); style_header_row(summary_sheet, 5)
            for item in status_breakdown: summary_sheet.append([f"  - {item['status'].title()}", item['count']])
            summary_sheet.append([])
            
            summary_sheet.append(['By Priority']); style_header_row(summary_sheet, summary_sheet.max_row)
            for item in priority_breakdown: summary_sheet.append([f"  - {item['priority'].title()}", item['count']])
            summary_sheet.append([])
            
            summary_sheet.append(['By Type']); style_header_row(summary_sheet, summary_sheet.max_row)
            for item in type_breakdown: summary_sheet.append([f"  - {item['type'].title()}", item['count']])

        adjust_column_widths(summary_sheet)
        
        raw_sheet = wb.create_sheet(title="All Complaints Data")
        
        raw_sheet.append(['Date', 'Student', 'Type', 'Priority', 'Status', 'Course', 'Teacher', 'Title', 'Description', 'Resolution Notes'])
        style_header_row(raw_sheet)

        for c in complaints:
            course_title = "N/A"
            teacher_name = "N/A"
            
            if c.enrollment:
                if c.enrollment.course:
                    course_title = c.enrollment.course.title
                if c.enrollment.schedule_slot and c.enrollment.schedule_slot.teacher:
                    teacher_name = c.enrollment.schedule_slot.teacher.get_full_name()

            raw_sheet.append([
                c.created_at.strftime('%Y-%m-%d'),
                c.student.get_full_name(),
                c.get_type_display(),
                c.get_priority_display(),
                c.get_status_display(),
                course_title,
                teacher_name,
                c.title,
                c.description,
                c.resolution_notes
            ])
            
        adjust_column_widths(raw_sheet)
        
        _save_report_file(report, wb, f"complaints_{start_date}_{end_date}")

    except Exception as e:
        _handle_task_failure(report_id, e)

def generate_schedule_slot_performance_report(report_id, schedule_slot_id):
    try:
        report = Report.objects.get(id=report_id); report.status = 'processing'; report.save()
        slot = ScheduleSlot.objects.get(id=schedule_slot_id)
        enrollments = Enrollment.objects.filter(schedule_slot=slot, status__in=['active', 'completed']).select_related('student')
        wb = openpyxl.Workbook(); sheet = wb.active; sheet.title = f"Performance - {slot.course.title[:20]}"
        sheet.append([f"Performance Report for: {slot.course.title}"]); sheet.append([f"Teacher: {slot.teacher.get_full_name()}"]); sheet.append([])
        
        headers = ['Student Name', 'Attendance %', 'Avg Homework Grade', 'Avg Quiz Score %', 'Overall Grade']; sheet.append(headers); style_header_row(sheet, 4)
        student_data = []
        for enr in enrollments:
            total_lessons = slot.lessons_in_lessons_app.filter(status='completed').count()
            present_count = Attendance.objects.filter(enrollment=enr, attendance='present').count()
            attendance_pct = (present_count / total_lessons * 100) if total_lessons > 0 else 0
            avg_hw_grade = HomeworkGrade.objects.filter(enrollment=enr).aggregate(avg=Avg('grade'))['avg'] or 0
            avg_quiz_score = QuizAttempt.objects.filter(user=enr.student, quiz__schedule_slot=slot, status='completed').aggregate(avg=Avg('score'))['avg'] or 0
            overall = (attendance_pct + avg_hw_grade + (avg_quiz_score or 0)) / 3
            student_data.append({'name': enr.get_student_name(), 'attendance': attendance_pct, 'homework': avg_hw_grade, 'quiz': avg_quiz_score or 0, 'overall': overall})
        
        student_data = sorted(student_data, key=lambda x: x['overall'], reverse=True)
        for data in student_data: sheet.append([data['name'], f"{data['attendance']:.2f}%", f"{data['homework']:.2f}", f"{data['quiz']:.2f}%", f"{data['overall']:.2f}"])
        sheet.append([]); sheet.append(['Top 3 Performers'])
        for d in student_data[:3]: sheet.append([f"  - {d['name']}", f"{d['overall']:.2f}"])
        sheet.append([]); sheet.append(['Bottom 3 Performers'])
        for d in student_data[-3:]: sheet.append([f"  - {d['name']}", f"{d['overall']:.2f}"])
        
        adjust_column_widths(sheet); _save_report_file(report, wb, f"performance_slot_{slot.id}")
    except Exception as e: _handle_task_failure(report_id, e)

def generate_student_performance_report(report_id, enrollment_id):
    try:
        report = Report.objects.get(id=report_id); report.status = 'processing'; report.save()
        enrollment = Enrollment.objects.get(id=enrollment_id); student = enrollment.student; slot = enrollment.schedule_slot
        wb = openpyxl.Workbook(); sheet = wb.active; sheet.title = "My Performance"
        
        sheet.append([f"Performance Report for {student.get_full_name()}"]); sheet.append([f"Course: {slot.course.title}"]); sheet.append([f"Date Generated: {timezone.now().strftime('%Y-%m-%d')}"]); sheet.append([])
        
        total_lessons = slot.lessons_in_lessons_app.filter(status='completed').count()
        present_count = Attendance.objects.filter(enrollment=enrollment, attendance='present').count()
        attendance_pct = (present_count / total_lessons * 100) if total_lessons > 0 else 100
        avg_hw_grade = HomeworkGrade.objects.filter(enrollment=enrollment).aggregate(avg=Avg('grade'))['avg'] or 0
        avg_quiz_score = QuizAttempt.objects.filter(user=student, quiz__schedule_slot=slot, status='completed').aggregate(avg=Avg('score'))['avg'] or 0
        overall_grade = (attendance_pct * 0.3) + (avg_hw_grade * 0.4) + ((avg_quiz_score or 0) * 0.3)
        
        sheet.append(['Category', 'Your Score / Result']); style_header_row(sheet, 5)
        data = [('Attendance', f"{attendance_pct:.2f}% ({present_count}/{total_lessons} lessons)"), ('Average Homework Grade', f"{avg_hw_grade:.2f} / 100"), ('Average Quiz Score', f"{avg_quiz_score:.2f}%"), ('Overall Weighted Grade', f"{overall_grade:.2f}%")]
        for row in data: sheet.append(row)
        adjust_column_widths(sheet)
        _save_report_file(report, wb, f"student_performance_{student.id}_{slot.id}")
    except Exception as e: _handle_task_failure(report_id, e)

# --- SCHEDULER-CALLABLE WRAPPERS ---
def schedule_monthly_financial_report():
    import django_rq
    end_date = timezone.now().date(); start_date = end_date - timezone.timedelta(days=30)
    admin_user = User.objects.filter(user_type='admin').first()
    report = Report.objects.create(report_type='financial_summary_period', requested_by=admin_user, status='pending', parameters={'start_date': str(start_date), 'end_date': str(end_date)})
    job = django_rq.get_queue('default').enqueue(generate_financial_report, report.id, start_date, end_date)
    report.job_id = job.id; report.save()

def schedule_monthly_statistical_report():
    import django_rq
    end_date = timezone.now().date(); start_date = end_date - timezone.timedelta(days=30)
    admin_user = User.objects.filter(user_type='admin').first()
    report = Report.objects.create(report_type='statistical_summary_period', requested_by=admin_user, status='pending', parameters={'start_date': str(start_date), 'end_date': str(end_date)})
    job = django_rq.get_queue('default').enqueue(generate_statistical_report, report.id, start_date, end_date)
    report.job_id = job.id; report.save()